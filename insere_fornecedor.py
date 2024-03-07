from openpyxl import load_workbook

# Busca pelos fornecedores na planilha referente e os adiciona na planilha principal

def copiar_dados(planilha_origem_path, planilha_destino_path, mapeamento_fornecedores, linha_inicial):
    wb_origem = load_workbook(filename=planilha_origem_path)
    ws_origem = wb_origem.active

    wb_destino = load_workbook(filename=planilha_destino_path)
    ws_destino = wb_destino.active

    for row_origem in ws_origem.iter_rows(min_row=linha_inicial, values_only=True):
        linha_destino = ws_destino.max_row + 1

        for col_origem, col_destino in mapeamento_fornecedores.items():
            index_col_origem = list(mapeamento_fornecedores.keys()).index(col_origem)

            valor_origem = row_origem[index_col_origem]

            ws_destino.cell(row=linha_destino, column=col_destino, value=valor_origem)

    wb_destino.save(planilha_destino_path)

    print("Dados copiados com sucesso.")

mapeamento_fornecedores = {
    'CODIGO': 1,
    'RAZAO': 2,
    'ENDERECO': 6,
    'CIDADE': 8,
    'ESTADO': 9,
    'CEP': 10,
    'CNPJ': 11
}

copiar_dados('C:\\Users\\User\\Documents\\Docs - Importacao - Cigam\\FORNECEDORES.xlsx', 'C:\\Users\\User\\Documents\\Docs - Importacao - Cigam\\Base Para NomeCompleto.xlsx', mapeamento_fornecedores, 2)
