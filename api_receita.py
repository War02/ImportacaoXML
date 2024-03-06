import openpyxl
import requests
import json

arquivo_excel = "C:\\Users\\User\\Documents\\Docs - Importacao - Cigam\\TESTE API RECEITA.xlsx"

coluna_cnpj = 2

url_base = "https://www.sintegraws.com.br/api/v1/execute-api.php"

token_acesso = "A20A7A55-96E1-4E70-9514-713AB9218308"

headers = {
    "Accept": "application/json"
}

respostas_api = []

wb = openpyxl.load_workbook(arquivo_excel)

sheet = wb.active

for row in range(2, sheet.max_row + 1):
    cnpj = sheet.cell(row=row, column=coluna_cnpj).value

    if cnpj:
        cnpj_formatado = cnpj.replace(".", "").replace("-", "")

        url = f"{url_base}?token={token_acesso}&cnpj={cnpj_formatado}&plugin=ST"

        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            resposta_json = response.json()

            respostas_api.append(resposta_json)

        else:
            print(f"Erro na requisição para o CNPJ {cnpj}: {response.status_code}")

with open("resultados.json", "w") as f:
    json.dump(respostas_api, f, indent=4)

print("Processo finalizado!")
