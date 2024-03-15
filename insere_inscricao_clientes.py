import os
import re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# Extrai os dados desejados do XML e os salva na planilha principal de importação
def extrair_informacao_xml(arquivo_xml):
    try:
        # Tentar analisar o arquivo XML
        tree = ET.parse(arquivo_xml)
        root = tree.getroot()

        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

        data = {
            'cnpj': find_text(root, ns, './/nfe:dest/nfe:CNPJ'),
            'ie': find_text(root, ns, './/nfe:dest/nfe:IE')
        }

        return data
    except ET.ParseError:
        return None

def find_text(root, namespace, path):
    try:
        return root.find(path, namespace).text
    except AttributeError:
        return None

def buscar_em_pasta(pasta):
    # Caminho para o arquivo Excel existente
    excel_path = '\\\\192.168.10.51\\Dados\\Departamental\\TI\\!CIGAM\\Migracao\\Dados Usados para Migração\\Docs - Importacao - Cigam\\CLIENTES SEM IE.xlsx'
    # Mapeamento de colunas
    column_mapping = {
        'IE': 'Inscricao'
    }
    # Carrega a planilha existente
    wb = load_workbook(filename=excel_path)
    ws = wb.active

    for pasta_atual, _, arquivos in os.walk(pasta):
        for arquivo in arquivos:
            # Verificar se o arquivo é um arquivo XML
            if arquivo.endswith('.xml'):
                # Construir o caminho completo do arquivo
                caminho_arquivo = os.path.join(pasta_atual, arquivo)
                # Extrair informações do arquivo XML
                data = extrair_informacao_xml(caminho_arquivo)
                if data:
                    # Procurar o CNPJ na planilha
                    cnpj = data['cnpj']
                    ie = data['ie']
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                        for cell in row:
                            if cell.value == cnpj:
                                # Se o CNPJ já existe na planilha, atualiza os dados
                                column_name = column_mapping.get('IE')
                                for row in ws.iter_rows(min_row=1, max_row=1):
                                    for cell_header in row:
                                        if cell_header.value == column_name:
                                            col_letter = cell_header.column_letter
                                            ws[f"{col_letter}{cell.row}"] = ie
                                            break
                                break
                    else:
                        # Se o CNPJ não existe na planilha, adiciona uma nova linha
                        next_row = ws.max_row + 1
                        ws[f"A{next_row}"] = cnpj
                        ws[f"B{next_row}"] = ie

    # Salva as alterações de volta para o arquivo Excel
    wb.save(excel_path)

    print("Dados dos arquivos XML adicionados com sucesso à planilha existente.")

# Percorre todas as pastas de 201901 até 202402
for ano in range(2019, 2024):
    for mes in range(1, 13):
        nome_pasta = f"{ano}{mes:02d}"
        pasta_atual = os.path.join('\\\\192.168.10.51\\Dados\\Compartilhados\\NFE\\nfe megatron\\sped\\xml\\04774509000142\\NFe', nome_pasta)
        if os.path.exists(pasta_atual):
            buscar_em_pasta(pasta_atual)