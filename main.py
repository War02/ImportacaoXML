import os
import re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook


def extrair_informacao_xml(arquivo_xml):
    try:
        # Tentar analisar o arquivo XML
        tree = ET.parse(arquivo_xml)
        root = tree.getroot()

        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

        data = {
            'xNome': find_text(root, ns, './/nfe:dest/nfe:xNome'),
            'fone': find_text(root, ns, './/nfe:dest/nfe:enderDest/nfe:fone'),
            'xLgr': find_text(root, ns, './/nfe:dest/nfe:enderDest/nfe:xLgr'),
            'xBairro': find_text(root, ns, './/nfe:dest/nfe:enderDest/nfe:xBairro'),
            'xMun': find_text(root, ns, './/nfe:dest/nfe:enderDest/nfe:xMun'),
            'uf': find_text(root, ns, './/nfe:dest/nfe:enderDest/nfe:UF'),
            'cep': find_text(root, ns, './/nfe:dest/nfe:enderDest/nfe:CEP'),
            'cnpj': find_text(root, ns, './/nfe:dest/nfe:CNPJ'),
            'cPais': find_text(root, ns, './/nfe:dest/nfe:enderDest/nfe:cPais'),
            'nro': find_text(root, ns, './/nfe:dest/nfe:enderDest/nfe:nro')
        }

        # Verifica se o endereço contém uma vírgula seguida de um número
        if ',' in data['xLgr']:
            numero_match = re.search(r',\s*(\d+)\s*$', data['xLgr'])
            if numero_match:
                data['nro'] = numero_match.group(1)

        return data
    except ET.ParseError:
        return None

def find_text(root, namespace, path):
    try:
        return root.find(path, namespace).text
    except AttributeError:
        return None

def buscar_em_pasta(pasta):
    # Caminho para o arquivo Excel existente
    excel_path = 'C:\\Users\\User\\Documents\\EMPRESAS_MODELOONDEPERMANECE O COD ANTERI(1).xlsx'
    # Mapeamento de colunas
    column_mapping = {
        'xNome': 'Nome Completo',
        'fone': 'Fone',
        'xLgr': 'Endereco',
        'xBairro': 'Bairro',
        'xMun': 'Municipio',
        'uf': 'UF',
        'cep': 'CEP',
        'cnpj': 'CNPJ_CPF',
        'cPais': 'Codigo Pais',
        'nro': 'Numero'
    }
    # Carrega a planilha existente
    wb = load_workbook(filename=excel_path)
    ws = wb.active

    for pasta_atual, _, arquivos in os.walk(pasta):
        for arquivo in arquivos:
            # Verificar se o arquivo é um arquivo XML
            if arquivo.endswith('.xml'):
                # Construir o caminho completo do arquivo
                caminho_arquivo = os.path.join(pasta_atual, arquivo)
                # Extrair informações do arquivo XML
                data = extrair_informacao_xml(caminho_arquivo)
                if data:
                    # Procurar o CNPJ na planilha
                    cnpj = data['cnpj']
                    for cell in ws['A']:
                        if cell.value == cnpj:
                            # Se o CNPJ já existe na planilha, atualiza os dados
                            for key, value in data.items():
                                column_name = column_mapping.get(key)
                                if column_name:
                                    for cell_header in ws[1]:
                                        if cell_header.value == column_name:
                                            col_letter = cell_header.column_letter
                                            ws[f"{col_letter}{cell.row}"] = value
                                            break
                            break
                    else:
                        # Se o CNPJ não existe na planilha, adiciona uma nova linha
                        next_row = ws.max_row + 1
                        for key, value in data.items():
                            column_name = column_mapping.get(key)
                            if column_name:
                                for cell_header in ws[1]:
                                    if cell_header.value == column_name:
                                        col_letter = cell_header.column_letter
                                        ws[f"{col_letter}{next_row}"] = value
                                        break

    # Salva as alterações de volta para o arquivo Excel
    wb.save(excel_path)

    print("Dados dos arquivos XML adicionados com sucesso à planilha existente.")

# Percorre todas as pastas de 201901 até 202402
for ano in range(2019, 2025):
    for mes in range(1, 13):
        nome_pasta = f"{ano}{mes:02d}"
        pasta_atual = os.path.join('\\\\192.168.10.51\\Dados\\Compartilhados\\NFE\\nfe megatron\\sped\\xml\\04774509000142\\NFe', nome_pasta)
        if os.path.exists(pasta_atual):
            buscar_em_pasta(pasta_atual)
